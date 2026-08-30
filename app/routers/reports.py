import csv
import io
from collections import Counter
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.dependencies import require_admin_roles
from app.models.entities import (
    Admin,
    Certificate,
    EventConfig,
    Fixture,
    Payment,
    Ped,
    Registration,
)

router = APIRouter(prefix="/admin/reports", tags=["Reports"])
REPORT_TYPES = {
    "registrations",
    "students",
    "payments",
    "attendance",
    "certificates",
    "fixtures",
}


def value(item):
    return item.isoformat() if isinstance(item, datetime) else item


async def load_registrations(
    db: AsyncSession,
    event_config_id: str | None,
    college: str | None,
    status: str | None,
    payment_status: str | None,
):
    query = (
        select(Registration)
        .options(
            selectinload(Registration.ped),
            selectinload(Registration.event_config),
            selectinload(Registration.students),
            selectinload(Registration.payments),
        )
        .order_by(Registration.created_at.desc())
    )
    if event_config_id:
        query = query.where(Registration.event_config_id == event_config_id)
    if college:
        query = query.where(Registration.college_name.ilike(f"%{college}%"))
    if status:
        query = query.where(Registration.status == status)
    if payment_status:
        query = query.where(Registration.payment_status == payment_status)
    return (await db.scalars(query)).unique().all()


async def latest_certificates(db: AsyncSession, registration_ids: list[str]):
    if not registration_ids:
        return {}
    rows = (
        await db.scalars(
            select(Certificate)
            .where(Certificate.registration_id.in_(registration_ids))
            .order_by(Certificate.student_id, Certificate.version.desc())
        )
    ).all()
    result = {}
    for certificate in rows:
        result.setdefault(certificate.student_id, certificate)
    return result


def build_rows(report_type, registrations, certificates):
    if report_type == "registrations":
        headers = [
            "Registration ID",
            "Registration Code",
            "College",
            "PED Name",
            "PED Email",
            "PED Contact",
            "Student Coordinator",
            "Coordinator Contact",
            "Sport",
            "Category",
            "Student Count",
            "Fee Paise",
            "Payment Status",
            "Approval Status",
            "Created At",
            "Submitted At",
            "Approved At",
            "Attendance Confirmed At",
            "Admin Note",
        ]
        rows = [
            [
                r.id,
                r.registration_code,
                r.college_name,
                r.ped.name,
                r.ped.official_email,
                r.ped_contact,
                r.student_coordinator_name,
                r.student_coordinator_contact,
                r.event_config.sport_name,
                r.event_config.category,
                len(r.students),
                r.fee_paise,
                r.payment_status,
                r.status,
                value(r.created_at),
                value(r.submitted_at),
                value(r.approved_at),
                value(r.attendance_confirmed_at),
                r.admin_note,
            ]
            for r in registrations
        ]
    elif report_type in {"students", "attendance"}:
        headers = [
            "Registration Code",
            "College",
            "PED Email",
            "Sport",
            "Category",
            "Student Name",
            "USN / Student ID",
            "Semester",
            "Contact",
            "Photo Uploaded",
            "Attendance Status",
            "Attendance Note",
            "Attendance Checked At",
            "Certificate Status",
            "Certificate Number",
        ]
        rows = []
        for r in registrations:
            for s in r.students:
                c = certificates.get(s.id)
                if report_type == "attendance" and s.attendance_status == "UNKNOWN":
                    continue
                rows.append(
                    [
                        r.registration_code,
                        r.college_name,
                        r.ped.official_email,
                        r.event_config.sport_name,
                        r.event_config.category,
                        s.full_name,
                        s.usn,
                        s.current_semester,
                        s.contact_number,
                        "YES" if s.photo_path else "NO",
                        s.attendance_status,
                        s.attendance_note,
                        value(s.attendance_checked_at),
                        c.status if c else "NOT_GENERATED",
                        c.certificate_number if c else "",
                    ]
                )
    elif report_type == "payments":
        headers = [
            "Registration Code",
            "College",
            "PED Email",
            "Sport",
            "Category",
            "Order ID",
            "Payment ID",
            "Amount Paise",
            "Currency",
            "Status",
            "Paid At",
            "Created At",
        ]
        rows = []
        for r in registrations:
            for p in r.payments:
                rows.append(
                    [
                        r.registration_code,
                        r.college_name,
                        r.ped.official_email,
                        r.event_config.sport_name,
                        r.event_config.category,
                        p.order_id,
                        p.payment_id,
                        p.amount_paise,
                        p.currency,
                        p.status,
                        value(p.paid_at),
                        value(p.created_at),
                    ]
                )
    elif report_type == "certificates":
        headers = [
            "Registration Code",
            "College",
            "PED Email",
            "Sport",
            "Category",
            "Student Name",
            "USN",
            "Certificate Number",
            "Version",
            "Status",
            "Published At",
            "Download Count",
            "Last Downloaded At",
        ]
        rows = []
        for r in registrations:
            for s in r.students:
                c = certificates.get(s.id)
                if c:
                    rows.append(
                        [
                            r.registration_code,
                            r.college_name,
                            r.ped.official_email,
                            r.event_config.sport_name,
                            r.event_config.category,
                            s.full_name,
                            s.usn,
                            c.certificate_number,
                            c.version,
                            c.status,
                            value(c.published_at),
                            c.download_count,
                            value(c.last_downloaded_at),
                        ]
                    )
    else:
        raise ValueError(report_type)
    return headers, rows


async def fixture_rows(db: AsyncSession, event_config_id: str | None):
    query = select(Fixture).options(selectinload(Fixture.event_config)).order_by(
        Fixture.created_at.desc()
    )
    if event_config_id:
        query = query.where(Fixture.event_config_id == event_config_id)
    fixtures = (await db.scalars(query)).all()
    headers = [
        "Fixture ID",
        "Sport",
        "Category",
        "Title",
        "Version",
        "Visibility",
        "Status",
        "Published At",
        "Supersedes ID",
        "File Path",
    ]
    rows = [
        [
            f.id,
            f.event_config.sport_name if f.event_config else "General",
            f.event_config.category if f.event_config else "All",
            f.title,
            f.version,
            f.visibility,
            f.status,
            value(f.published_at),
            f.supersedes_id,
            f.file_path,
        ]
        for f in fixtures
    ]
    return headers, rows


@router.get("/export.csv")
async def export_csv(
    report: str = Query("students"),
    event_config_id: str | None = None,
    college: str | None = None,
    status: str | None = None,
    payment_status: str | None = None,
    admin: Admin = Depends(require_admin_roles("REGISTRATION_ADMIN")),
    db: AsyncSession = Depends(get_db),
):
    if report not in REPORT_TYPES:
        raise HTTPException(
            status_code=422, detail=f"report must be one of {sorted(REPORT_TYPES)}"
        )
    if report == "fixtures":
        headers, rows = await fixture_rows(db, event_config_id)
    else:
        registrations = await load_registrations(
            db, event_config_id, college, status, payment_status
        )
        certs = await latest_certificates(db, [r.id for r in registrations])
        headers, rows = build_rows(report, registrations, certs)
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    writer.writerows(rows)
    binary = io.BytesIO(stream.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        binary,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=bnm-{report}-report.csv"
        },
    )


@router.get("/students.csv", include_in_schema=False)
async def legacy_students_csv(
    event_config_id: str | None = None,
    college: str | None = None,
    admin: Admin = Depends(require_admin_roles("REGISTRATION_ADMIN")),
    db: AsyncSession = Depends(get_db),
):
    return await export_csv(
        "students", event_config_id, college, None, None, admin, db
    )


def style_sheet(sheet):
    fill = PatternFill("solid", fgColor="17324D")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        sheet.column_dimensions[column[0].column_letter].width = width


@router.get("/workbook.xlsx")
async def workbook(
    event_config_id: str | None = None,
    college: str | None = None,
    status: str | None = None,
    payment_status: str | None = None,
    admin: Admin = Depends(require_admin_roles("REGISTRATION_ADMIN")),
    db: AsyncSession = Depends(get_db),
):
    registrations = await load_registrations(
        db, event_config_id, college, status, payment_status
    )
    certs = await latest_certificates(db, [r.id for r in registrations])
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Summary")
    status_counts = Counter(r.status for r in registrations)
    payment_counts = Counter(r.payment_status for r in registrations)
    summary.append(["Metric", "Value"])
    summary.append(["Registrations", len(registrations)])
    summary.append(["Students", sum(len(r.students) for r in registrations)])
    for key, count in sorted(status_counts.items()):
        summary.append([f"Approval - {key}", count])
    for key, count in sorted(payment_counts.items()):
        summary.append([f"Payment - {key}", count])
    summary.append(
        [
            "Present Students",
            sum(
                1
                for r in registrations
                for student in r.students
                if student.attendance_status == "PRESENT"
            ),
        ]
    )
    summary.append(["Certificates", len(certs)])
    style_sheet(summary)

    for report_type, title in [
        ("registrations", "Registrations"),
        ("students", "Students"),
        ("payments", "Payments"),
        ("attendance", "Attendance"),
        ("certificates", "Certificates"),
    ]:
        headers, rows = build_rows(report_type, registrations, certs)
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        style_sheet(sheet)

    headers, rows = await fixture_rows(db, event_config_id)
    fixture_sheet = workbook.create_sheet("Fixtures")
    fixture_sheet.append(headers)
    for row in rows:
        fixture_sheet.append(row)
    style_sheet(fixture_sheet)

    binary = io.BytesIO()
    workbook.save(binary)
    binary.seek(0)
    return StreamingResponse(
        binary,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=bnmit-odyssey-report.xlsx"
        },
    )
